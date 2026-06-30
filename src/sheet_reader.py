from __future__ import annotations

import re
import os
import tempfile
import time
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


SOCIAL_DOMAINS = ("threads.", "instagram.", "facebook.", "fb.watch", "fb.com")
DEFAULT_URL_COLUMN = os.getenv("THREADS_SHEET_URL_COLUMN", "B")
FALLBACK_SCAN_ENABLED = os.getenv("THREADS_SHEET_SCAN_FALLBACK", "true").strip().lower() in {"1", "true", "yes", "y", "on"}
SHEET_CACHE_SECONDS = int(os.getenv("THREADS_SHEET_CACHE_SECONDS", "600") or "0")


@dataclass(frozen=True)
class SheetUrlRow:
    row_number: int
    platform: str
    url: str
    title: str
    author: str
    published_at: str
    source_column: str


def load_sheet_urls(sheet_source: str | Path, *, platforms: set[str] | None = None) -> list[SheetUrlRow]:
    rows, _metadata = inspect_sheet_urls(sheet_source, platforms=platforms)
    return rows


def inspect_sheet_urls(sheet_source: str | Path, *, platforms: set[str] | None = None) -> tuple[list[SheetUrlRow], dict[str, object]]:
    workbook_path = _resolve_workbook_path(sheet_source)
    workbook = load_workbook(workbook_path, data_only=False)
    worksheet = workbook.active
    header_row, headers, column_indexes = _find_header_row(worksheet)

    rows: list[SheetUrlRow] = []
    seen: set[tuple[str, str]] = set()
    normalized_platforms = {platform.upper() for platform in platforms} if platforms else None

    for row_number in range(header_row + 1, worksheet.max_row + 1):
        platform = _normalize_platform(worksheet.cell(row_number, column_indexes["source"]).value)
        if not platform or platform == "來源":
            continue
        if normalized_platforms and platform not in normalized_platforms:
            continue

        url, source_column = _extract_best_url(worksheet, headers, row_number, column_indexes)
        if not url:
            continue
        key = (platform, url)
        if key in seen:
            continue
        seen.add(key)

        rows.append(
            SheetUrlRow(
                row_number=row_number,
                platform=platform,
                url=url,
                title=str(worksheet.cell(row_number, column_indexes["title"]).value or "").strip(),
                author=str(worksheet.cell(row_number, column_indexes["author"]).value or "").strip(),
                published_at=_cell_text(worksheet.cell(row_number, column_indexes["time"])) if "time" in column_indexes else "",
                source_column=source_column,
            )
        )
    metadata = {
        "sheet_name": worksheet.title,
        "header_row": header_row,
        "headers": headers,
        "columns": {
            key: headers[index - 1] if index - 1 < len(headers) else f"column_{index}"
            for key, index in column_indexes.items()
        },
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
    }
    return rows, metadata


def _resolve_workbook_path(sheet_source: str | Path) -> Path:
    source = str(sheet_source)
    local_path = Path(source).expanduser()
    if local_path.exists():
        return local_path
    if source.startswith(("http://", "https://")):
        return _download_google_sheet(source)
    raise FileNotFoundError(f"Sheet source not found: {sheet_source}")


def _download_google_sheet(url: str) -> Path:
    spreadsheet_id = _extract_spreadsheet_id(url)
    if not spreadsheet_id:
        raise ValueError(f"Cannot extract Google spreadsheet id from URL: {url}")

    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    gid = _extract_gid(url)
    if gid:
        export_url += f"&gid={urllib.parse.quote(gid)}"

    cache_suffix = f"_{gid}" if gid else ""
    target = Path(tempfile.gettempdir()) / f"threads_sheet_{spreadsheet_id}{cache_suffix}.xlsx"
    if _cached_file_is_fresh(target):
        return target

    request = urllib.request.Request(
        export_url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
            )
        },
    )
    with urllib.request.urlopen(request, timeout=60) as response:
        content_type = response.headers.get("content-type", "")
        data = response.read()
    if not data.startswith(b"PK"):
        raise RuntimeError(
            "Google Sheet export did not return an xlsx file. "
            f"Content-Type: {content_type}. Make sure the sheet link is accessible."
        )
    target.write_bytes(data)
    return target


def _extract_spreadsheet_id(url: str) -> str:
    match = re.search(r"/spreadsheets/d/([^/?#]+)", url)
    if match:
        return match.group(1)
    query = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    return query.get("id", [""])[0]


def _extract_gid(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    query = urllib.parse.parse_qs(parsed.query)
    gid = query.get("gid", [""])[0]
    if gid:
        return gid
    fragment_query = urllib.parse.parse_qs(parsed.fragment)
    return fragment_query.get("gid", [""])[0]


def _cached_file_is_fresh(path: Path) -> bool:
    if SHEET_CACHE_SECONDS <= 0 or not path.exists() or path.stat().st_size == 0:
        return False
    age = time.time() - path.stat().st_mtime
    return age <= SHEET_CACHE_SECONDS


def _find_columns(headers: list[str]) -> dict[str, int]:
    required = {
        "time": ("時間", "發布時間", "date", "time"),
        "title": ("文章標題", "標題", "fb標題", "fb標題_2", "title"),
        "author": ("作者", "頻道", "author"),
        "source": ("來源", "網站", "平台", "source"),
        "url": ("網址", "url", "URL"),
    }
    result: dict[str, int] = {}
    for key, candidates in required.items():
        for index, header in enumerate(headers, start=1):
            if header in candidates:
                result[key] = index
                break
    missing = [key for key in ("title", "source") if key not in result]
    if missing:
        raise ValueError(f"Sheet is missing required columns: {', '.join(missing)}")
    return result


def _find_header_row(worksheet) -> tuple[int, list[str], dict[str, int]]:
    for row_number in range(1, min(worksheet.max_row, 20) + 1):
        headers = [str(cell.value or "").strip() for cell in worksheet[row_number]]
        try:
            column_indexes = _find_columns(headers)
        except ValueError:
            continue
        return row_number, headers, column_indexes
    raise ValueError("Sheet is missing required columns: title, source")


def _normalize_platform(value: object) -> str:
    text = str(value or "").strip()
    lowered = text.lower()
    if not text:
        return ""
    if "facebook" in lowered or text in {"FB", "臉書", "Facebook粉絲團"}:
        return "FACEBOOK"
    if "instagram" in lowered or lowered == "ig":
        return "IG"
    if "threads" in lowered or "thread" in lowered or "串" in text:
        return "THREADS"
    return text.upper()


def _cell_text(cell) -> str:
    value = cell.value
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value or "").strip()


def _extract_best_url(worksheet, headers: list[str], row_number: int, column_indexes: dict[str, int]) -> tuple[str, str]:
    preferred_col = _column_ref_to_index(DEFAULT_URL_COLUMN)
    if preferred_col:
        preferred_cell = worksheet.cell(row_number, preferred_col)
        preferred_url = _cell_url(preferred_cell)
        if _is_social_url(preferred_url):
            header = headers[preferred_col - 1] if preferred_col - 1 < len(headers) else f"column_{preferred_col}"
            return preferred_url, header or DEFAULT_URL_COLUMN.upper()

    title_col = column_indexes["title"]
    if title_col != preferred_col:
        title_cell = worksheet.cell(row_number, title_col)
        title_url = _cell_url(title_cell)
        if _is_social_url(title_url):
            return title_url, headers[title_col - 1] or f"column_{title_col}"

    url_col = column_indexes.get("url")
    if url_col and url_col != preferred_col:
        url_cell = worksheet.cell(row_number, url_col)
        explicit_url = _cell_url(url_cell)
        if _is_social_url(explicit_url):
            return explicit_url, headers[url_col - 1] or f"column_{url_col}"

    if not FALLBACK_SCAN_ENABLED:
        return "", ""

    for column in range(1, worksheet.max_column + 1):
        if column in {preferred_col, title_col, url_col}:
            continue
        cell_url = _cell_url(worksheet.cell(row_number, column))
        if _is_social_url(cell_url):
            return cell_url, headers[column - 1] or f"column_{column}"
    return "", ""


def _column_ref_to_index(value: str) -> int | None:
    text = str(value or "").strip().upper()
    if not text:
        return None
    if text.isdigit():
        return max(1, int(text))
    if not re.fullmatch(r"[A-Z]+", text):
        return None
    index = 0
    for char in text:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index


def _cell_url(cell) -> str:
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    value = str(cell.value or "").strip()
    match = re.search(r"https?://\S+", value)
    return match.group(0).strip() if match else ""


def _is_social_url(url: str) -> bool:
    lowered = url.lower()
    return lowered.startswith(("http://", "https://")) and any(domain in lowered for domain in SOCIAL_DOMAINS)
