from __future__ import annotations

import argparse
import csv
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from exporter import (  # noqa: E402
    FIELDNAMES,
    REPORT_FIELDNAMES,
    append_raw_row,
    append_report_row,
    ensure_csv_schema,
    read_terminal_urls,
    write_csv,
)
from main import (  # noqa: E402
    InputItem,
    MISSING_PROFILE_CACHE_VALUE,
    apply_sheet_metadata,
    alternate_threads_domain,
    enrich_social_followers,
    facebook_profile_url,
    load_input_items,
    platform_from_url,
    profile_cache_key,
    read_urls,
    to_report_row,
)
from parser import (  # noqa: E402
    STATUS_LOGIN_REQUIRED,
    STATUS_POST_NOT_LOADED,
    STATUS_SUCCESS,
    parse_threads_page,
)
from scraper import (  # noqa: E402
    _looks_like_threads_data_response,
    _parse_jsonish_response_text,
    _post_id_from_url,
    _profile_url_from_post_url,
    _safe_slug,
)
from sheet_reader import inspect_sheet_urls  # noqa: E402
from social_parser import parse_profile_follower_count, parse_social_page  # noqa: E402


class ThreadsParserTests(unittest.TestCase):
    def test_parse_threads_visible_post_metrics_and_report_fields(self) -> None:
        post_url = "https://www.threads.com/@alice/post/ABC123"
        html = """
        <html><head>
          <meta property="og:url" content="https://www.threads.com/@alice/post/ABC123">
          <time datetime="2026-06-01T10:00:00Z"></time>
        </head><body></body></html>
        """
        visible_text = """
        alice
        2026-06-01
        New product update is live.
        Translate
        1.2K
        34
        5
        2
        """

        parsed = parse_threads_page(post_url, html, visible_text)

        self.assertEqual(parsed.status, STATUS_SUCCESS)
        self.assertEqual(parsed.row["username"], "@alice")
        self.assertEqual(parsed.row["text"], "New product update is live.")
        self.assertEqual(parsed.row["created_at"], "2026-06-01T10:00:00Z")
        self.assertEqual(parsed.row["like_count"], 1200)
        self.assertEqual(parsed.row["reply_count"], 34)
        self.assertEqual(parsed.row["repost_count"], 5)
        self.assertEqual(parsed.row["quote_count"], 2)

        raw_row = dict(parsed.row)
        self.assertNotIn("interaction_total", raw_row)
        self.assertNotIn("thread_weight", raw_row)
        self.assertNotIn("like_weight", raw_row)
        self.assertNotIn("share_weight", raw_row)
        self.assertNotIn("value", raw_row)

        report_row = to_report_row(1, raw_row, auth_mode="public", platform="THREADS")
        self.assertEqual(report_row["FB"], "Threads > @alice")
        self.assertEqual(report_row["分享"], 7)
        self.assertEqual(report_row["reach_status"], "public_unavailable")
        self.assertNotIn("互動總次數", report_row)
        self.assertNotIn("VALUE", report_row)
        self.assertNotIn("interaction_total", FIELDNAMES)
        self.assertNotIn("thread_weight", FIELDNAMES)
        self.assertNotIn("like_weight", FIELDNAMES)
        self.assertNotIn("share_weight", FIELDNAMES)
        self.assertNotIn("value", FIELDNAMES)
        self.assertNotIn("VALUE", REPORT_FIELDNAMES)
        self.assertNotIn("status", REPORT_FIELDNAMES)
        self.assertNotIn("reach_status", REPORT_FIELDNAMES)

    def test_threads_login_and_profile_redirect_statuses_are_terminal_signals(self) -> None:
        login = parse_threads_page(
            "https://www.threads.com/@alice/post/ABC123",
            "<html><body>Log in to Threads Continue with Instagram</body></html>",
            "Log in to Threads\nContinue with Instagram",
        )
        self.assertEqual(login.status, STATUS_LOGIN_REQUIRED)

        redirected_profile = parse_threads_page(
            "https://www.threads.com/@alice/post/ABC123",
            """
            <html><head>
              <meta property="og:url" content="https://www.threads.com/@alice">
            </head><body></body></html>
            """,
            "alice\n提及\n串文\n回覆\n影音內容",
        )
        self.assertEqual(redirected_profile.status, STATUS_POST_NOT_LOADED)

    def test_parse_threads_relative_time_visible_metrics(self) -> None:
        post_url = "https://www.threads.com/@example_travel/post/RELATIVE123"
        html = '<html><head><meta property="og:url" content="https://www.threads.com/@example_travel/post/RELATIVE123"></head></html>'
        visible_text = """
        串文
        1.6 萬次瀏覽
        example_travel
        Travel sample
        1天
        Synthetic travel post used for parser regression testing.
        翻譯
        141
        17
        2
        7
        example_reply
        1天
        Synthetic reply text.
        """

        parsed = parse_threads_page(
            post_url,
            html,
            visible_text,
            [{"reply_count": 0, "like_count": 141, "code": "RELATIVE123"}],
        )

        self.assertEqual(parsed.status, STATUS_SUCCESS)
        self.assertEqual(parsed.row["like_count"], 141)
        self.assertEqual(parsed.row["reply_count"], 17)
        self.assertEqual(parsed.row["repost_count"], 2)
        self.assertEqual(parsed.row["quote_count"], 7)
        self.assertEqual(parsed.row["view_count"], 16000)

    def test_threads_post_metrics_ignore_related_thread_json(self) -> None:
        post_url = "https://www.threads.com/@ke_song_tw/post/DaNC74PlEoV"
        html = '<html><head><meta property="og:url" content="https://www.threads.com/@ke_song_tw/post/DaNC74PlEoV"></head></html>'
        visible_text = """
        串文
        144次瀏覽
        ke_song_tw
        1小時
        2026世界盃開吃計畫🔥
        翻譯
        相關串文
        ke_song_tw
        3天
        Related post text.
        翻譯
        12
        6
        4
        2
        """

        parsed = parse_threads_page(
            post_url,
            html,
            visible_text,
            [
                {
                    "shortcode": "DaNC74PlEoV",
                    "related_threads": [
                        {
                            "shortcode": "RELATED_IN_CONTAINER",
                            "like_count": 0,
                            "text_post_app_info": {"direct_reply_count": 0, "repost_count": 0, "quote_count": 0},
                        }
                    ],
                },
                {
                    "shortcode": "RELATED123",
                    "like_count": 99,
                    "text_post_app_info": {
                        "direct_reply_count": 8,
                        "repost_count": 7,
                        "quote_count": 6,
                    },
                },
            ],
        )

        self.assertEqual(parsed.status, STATUS_SUCCESS)
        self.assertEqual(parsed.row["like_count"], "N/A")
        self.assertEqual(parsed.row["reply_count"], "N/A")
        self.assertEqual(parsed.row["repost_count"], "N/A")
        self.assertEqual(parsed.row["quote_count"], "N/A")
        self.assertEqual(parsed.row["view_count"], 144)

    def test_threads_post_metrics_use_matching_post_json(self) -> None:
        post_url = "https://www.threads.com/@alice/post/TARGET123"
        html = '<html><head><meta property="og:url" content="https://www.threads.com/@alice/post/TARGET123"></head></html>'
        visible_text = """
        alice
        1天
        Target post.
        Translate
        """

        parsed = parse_threads_page(
            post_url,
            html,
            visible_text,
            [
                {"shortcode": "RELATED123", "like_count": 99},
                {
                    "shortcode": "TARGET123",
                    "like_count": 7,
                    "text_post_app_info": {
                        "direct_reply_count": 2,
                        "repost_count": 1,
                        "quote_count": 0,
                    },
                },
            ],
        )

        self.assertEqual(parsed.status, STATUS_SUCCESS)
        self.assertEqual(parsed.row["like_count"], 7)
        self.assertEqual(parsed.row["reply_count"], 2)
        self.assertEqual(parsed.row["repost_count"], 1)
        self.assertEqual(parsed.row["quote_count"], 0)

    def test_threads_ignores_incomplete_visible_metric_group_from_image_ocr(self) -> None:
        post_url = "https://www.threads.com/@example/post/OCR123"
        html = '<html><head><meta property="og:url" content="https://www.threads.com/@example/post/OCR123"></head></html>'
        visible_text = """
        example
        1天
        Image post with numeric text.
        翻譯
        5
        1
        儲存
        另存新檔
        分享
        """

        parsed = parse_threads_page(post_url, html, visible_text)

        self.assertEqual(parsed.status, STATUS_SUCCESS)
        self.assertEqual(parsed.row["like_count"], "N/A")
        self.assertEqual(parsed.row["reply_count"], "N/A")
        self.assertEqual(parsed.row["repost_count"], "N/A")
        self.assertEqual(parsed.row["quote_count"], "N/A")


class SocialParserTests(unittest.TestCase):
    def test_parse_instagram_and_facebook_public_pages(self) -> None:
        ig_row = parse_social_page(
            "IG",
            "https://www.instagram.com/reel/REEL123/",
            """
            <html><head>
              <meta property="og:title" content="alice on Instagram">
              <meta property="og:description" content="A launch reel">
              <meta property="article:published_time" content="2026-06-02T08:00:00Z">
            </head>
            <body><span aria-label="1.5K likes"></span><span aria-label="42 comments"></span><span aria-label="9K views"></span></body></html>
            """,
            "1.5K likes\n42 comments\n9K views",
        )
        self.assertEqual(ig_row["status"], "success")
        self.assertEqual(ig_row["username"], "@alice")
        self.assertEqual(ig_row["like_count"], 1500)
        self.assertEqual(ig_row["reply_count"], 42)
        self.assertEqual(ig_row["view_count"], 9000)

        fb_row = parse_social_page(
            "FACEBOOK",
            "https://www.facebook.com/page/posts/123",
            """
            <html><head>
              <meta property="og:title" content="Example Page">
              <meta property="og:description" content="Campaign post">
            </head>
            <body><span aria-label="所有心情: 88"></span><span aria-label="7 comments"></span><span aria-label="3 shares"></span></body></html>
            """,
            "88 likes\n7 comments\n3 shares",
        )
        self.assertEqual(fb_row["status"], "success")
        self.assertEqual(fb_row["username"], "Example Page")
        self.assertEqual(fb_row["like_count"], 88)
        self.assertEqual(fb_row["reply_count"], 7)
        self.assertEqual(fb_row["repost_count"], 3)

    def test_parse_facebook_reel_rail_metrics_from_bare_counts(self) -> None:
        fb_row = parse_social_page(
            "FACEBOOK",
            "https://www.facebook.com/reel/123",
            """
            <html><head>
              <meta property="og:title" content="Example Page">
              <meta property="og:description" content="Campaign reel">
            </head><body></body></html>
            """,
            """
            Campaign reel
            5
            1
            9
            儲存
            另存新檔
            分享
            傳送至Keep筆記
            """,
        )

        self.assertEqual(fb_row["status"], "success")
        self.assertEqual(fb_row["like_count"], 5)
        self.assertEqual(fb_row["reply_count"], 1)
        self.assertEqual(fb_row["repost_count"], 9)

    def test_parse_facebook_reel_label_from_regular_post_url(self) -> None:
        fb_row = parse_social_page(
            "FACEBOOK",
            "https://www.facebook.com/286603368359862_1449172157254240",
            """
            <html><head>
              <meta property="og:title" content="Example Page">
              <meta property="og:description" content="Campaign reel">
            </head><body></body></html>
            """,
            """
            Campaign reel
            Example Page
            5
            1
            Reel
            在 Facebook 查看更多
            """,
        )

        self.assertEqual(fb_row["status"], "success")
        self.assertEqual(fb_row["like_count"], 5)
        self.assertEqual(fb_row["reply_count"], 1)
        self.assertEqual(fb_row["repost_count"], "N/A")

    def test_profile_follower_count_from_text_and_json(self) -> None:
        self.assertEqual(parse_profile_follower_count("IG", "", "1.2萬 followers"), 12000)
        self.assertEqual(parse_profile_follower_count("FACEBOOK", "", "1.5 億位追蹤者"), 150000000)
        self.assertEqual(parse_profile_follower_count("FACEBOOK", "", "3.4萬人追蹤"), 34000)
        self.assertEqual(parse_profile_follower_count("FACEBOOK", "", "2.5M followers"), 2500000)
        self.assertEqual(parse_profile_follower_count("FACEBOOK", "", ".\n粉絲專頁\n7.5 萬位追蹤者"), 75000)
        self.assertEqual(
            parse_profile_follower_count("FACEBOOK", "", "", [{"profile": {"followerCount": {"count": "3K"}}}]),
            3000,
        )


class SheetAndInputTests(unittest.TestCase):
    def _save_workbook(self, workbook: Workbook, directory: str, name: str) -> Path:
        path = Path(directory) / name
        workbook.save(path)
        return path

    def test_sheet_reader_uses_hyperlink_platform_filtering_and_deduplication(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "sample.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["時間", "文章標題", "作者", "來源", "網址"])
            ws.append(["2026-06-01", "Thread title", "@alice", "Threads", ""])
            ws["B2"].hyperlink = "https://www.threads.com/@alice/post/ABC123"
            ws.append(["2026-06-02", "IG title", "@bob", "IG", "https://www.instagram.com/p/IG123/"])
            ws.append(["2026-06-02", "IG title duplicate", "@bob", "IG", "https://www.instagram.com/p/IG123/"])
            ws.append(["2026-06-03", "FB title", "Page", "Facebook", "https://www.facebook.com/page/posts/1"])
            wb.save(workbook_path)

            rows, metadata = inspect_sheet_urls(workbook_path, platforms={"THREADS", "IG"})

        self.assertEqual(metadata["header_row"], 1)
        self.assertEqual([row.platform for row in rows], ["THREADS", "IG"])
        self.assertEqual(rows[0].url, "https://www.threads.com/@alice/post/ABC123")
        self.assertEqual(rows[0].source_column, "文章標題")
        self.assertEqual(rows[1].url, "https://www.instagram.com/p/IG123/")

    def test_sheet_reader_accepts_repeated_google_sheet_layout_variants(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            hyperlink_wb = Workbook()
            ws = hyperlink_wb.active
            ws.append(["匯出日期", "2026-06-01"])
            ws.append(["備註", "header starts below"])
            ws.append(["時間", "文章標題", "作者", "來源", "網址"])
            ws.append(["2026-06-01", "Thread title", "@alice", "串文", ""])
            ws["B4"].hyperlink = "https://www.threads.com/@alice/post/ABC123"
            ws.append(["2026-06-02", "IG title", "@bob", "instagram", ""])
            ws["B5"].hyperlink = "https://www.instagram.com/p/IG123/"
            hyperlink_path = self._save_workbook(hyperlink_wb, tmp, "header_offset_hyperlinks.xlsx")

            explicit_url_wb = Workbook()
            ws = explicit_url_wb.active
            ws.append(["時間", "文章標題", "作者", "來源", "網址", "備註"])
            ws.append(["2026-06-01", "Thread title", "@alice", "Threads", "https://www.threads.com/@alice/post/ABC123", ""])
            ws.append(["2026-06-02", "FB title", "Page", "臉書", "https://www.facebook.com/page/posts/1", ""])
            explicit_url_path = self._save_workbook(explicit_url_wb, tmp, "explicit_url_column.xlsx")

            fallback_wb = Workbook()
            ws = fallback_wb.active
            ws.append(["時間", "文章標題", "作者", "來源", "備用連結"])
            ws.append(["2026-06-01", "No hyperlink title", "@alice", "THREADS", "https://www.threads.com/@alice/post/ABC123"])
            fallback_path = self._save_workbook(fallback_wb, tmp, "fallback_scan.xlsx")

            hyperlink_rows, hyperlink_meta = inspect_sheet_urls(hyperlink_path, platforms={"THREADS", "IG", "FACEBOOK"})
            explicit_rows, _ = inspect_sheet_urls(explicit_url_path, platforms={"THREADS", "FACEBOOK"})
            fallback_rows, _ = inspect_sheet_urls(fallback_path, platforms={"THREADS"})

        self.assertEqual(hyperlink_meta["header_row"], 3)
        self.assertEqual([row.platform for row in hyperlink_rows], ["THREADS", "IG"])
        self.assertEqual(hyperlink_rows[0].source_column, "文章標題")
        self.assertEqual([row.platform for row in explicit_rows], ["THREADS", "FACEBOOK"])
        self.assertEqual(explicit_rows[0].source_column, "網址")
        self.assertEqual(len(fallback_rows), 1)
        self.assertEqual(fallback_rows[0].source_column, "備用連結")

    def test_read_urls_and_load_input_items_are_deduplicated_and_platform_aware(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            input_path = Path(tmp) / "urls.txt"
            input_path.write_text(
                "\n".join(
                    [
                        "# comment",
                        "https://www.threads.com/@alice/post/ABC123",
                        "https://www.threads.com/@alice/post/ABC123",
                        "https://www.instagram.com/p/IG123/",
                        "https://fb.watch/FB123",
                    ]
                ),
                encoding="utf-8",
            )
            urls = read_urls(input_path)
            args = argparse.Namespace(input=str(input_path), sheet="", sheet_platforms="THREADS")
            items, label = load_input_items(args)

        self.assertEqual(urls, [
            "https://www.threads.com/@alice/post/ABC123",
            "https://www.instagram.com/p/IG123/",
            "https://fb.watch/FB123",
        ])
        self.assertEqual(label, str(input_path))
        self.assertEqual([item.platform for item in items], ["THREADS", "IG", "FACEBOOK"])


class ExporterAndMainHelperTests(unittest.TestCase):
    def test_csv_append_read_and_schema_backup_behaviors(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "metrics.csv"
            write_csv([], output_path)
            append_raw_row(
                {
                    "post_url": "https://www.threads.com/@alice/post/ABC123",
                    "username": "@alice",
                    "status": "success",
                },
                output_path,
            )
            append_raw_row(
                {
                    "post_url": "https://www.threads.com/@alice/post/MISSING",
                    "username": "@alice",
                    "status": "post_not_loaded",
                },
                output_path,
            )

            self.assertEqual(
                read_terminal_urls(output_path, report_format=False),
                {
                    "https://www.threads.com/@alice/post/ABC123",
                    "https://www.threads.com/@alice/post/MISSING",
                },
            )

            with output_path.open(newline="", encoding="utf-8-sig") as csvfile:
                header = next(csv.reader(csvfile))
            self.assertEqual(header, FIELDNAMES)

            backup = ensure_csv_schema(output_path, ["different"])
            self.assertIsNotNone(backup)
            self.assertTrue(backup.exists())
            self.assertFalse(output_path.exists())

    def test_reusing_same_output_filename_rewrites_cleanly_and_appends_one_header(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "same_name.csv"
            first_row = {
                "post_url": "https://www.threads.com/@alice/post/OLD",
                "username": "@alice",
                "status": "success",
            }
            second_row = {
                "post_url": "https://www.threads.com/@alice/post/NEW",
                "username": "@alice",
                "status": "success",
            }

            write_csv([first_row], output_path)
            write_csv([], output_path)
            append_raw_row(second_row, output_path)

            with output_path.open(newline="", encoding="utf-8-sig") as csvfile:
                rows = list(csv.reader(csvfile))

        self.assertEqual(rows[0], FIELDNAMES)
        self.assertEqual(len(rows), 2)
        self.assertIn("NEW", rows[1][FIELDNAMES.index("post_url")])
        self.assertNotIn("OLD", "\n".join(",".join(row) for row in rows))

    def test_report_csv_omits_internal_value_and_status_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "report.csv"
            append_report_row(
                {
                    "序號": 1,
                    "發布時間": "2026-06-01",
                    "FB": "Threads > @alice",
                    "fb標題": "Title",
                    "網站": "Threads",
                    "頻道": "@alice",
                    "fb標題_2": "Title",
                    "討論串總則數": 1,
                    "點閱數/按讚數": 2,
                    "瀏覽數": "N/A",
                    "分享": 3,
                    "網址": "https://www.threads.com/@alice/post/ABC123",
                    "粉絲團追蹤人數": "N/A",
                    "觸及": "N/A",
                    "status": "success",
                    "reach_status": "public_unavailable",
                },
                output_path,
            )

            with output_path.open(newline="", encoding="utf-8-sig") as csvfile:
                header = next(csv.reader(csvfile))

        self.assertEqual(header, REPORT_FIELDNAMES)
        self.assertEqual(
            header,
            ["網址", "fb標題", "討論串總則數", "點閱數/按讚數", "瀏覽數", "分享", "粉絲團追蹤人數", "觸及"],
        )
        self.assertNotIn("討論串加權(*45)", header)
        self.assertNotIn("讚數加權(*30)", header)
        self.assertNotIn("分享加權(*60)", header)
        self.assertNotIn("互動總次數", header)
        self.assertNotIn("VALUE", header)
        self.assertNotIn("status", header)
        self.assertNotIn("reach_status", header)

    def test_main_helpers_preserve_report_metadata_and_url_conventions(self) -> None:
        row = {
            "post_url": "https://www.instagram.com/p/IG123/",
            "username": "N/A",
            "text": "N/A",
            "created_at": "N/A",
            "like_count": 10,
            "reply_count": 2,
            "repost_count": 0,
            "quote_count": 0,
            "view_count": "N/A",
            "follower_count": "N/A",
            "reach": "N/A",
            "status": "success",
        }
        item = InputItem(
            url=row["post_url"],
            platform="IG",
            title="Sheet title",
            author="@sheet_author",
            published_at="2026-06-01",
        )
        apply_sheet_metadata(row, item)
        report_row = to_report_row(2, row, auth_mode="session_state", platform="IG")

        self.assertEqual(row["username"], "@sheet_author")
        self.assertEqual(row["text"], "Sheet title")
        self.assertEqual(row["created_at"], "2026-06-01")
        self.assertEqual(report_row["FB"], "Instagram > @sheet_author")
        self.assertEqual(report_row["reach_status"], "unavailable_after_login")
        self.assertEqual(platform_from_url("https://fb.watch/ABC"), "FACEBOOK")
        self.assertEqual(alternate_threads_domain("https://www.threads.com/@a/post/1"), "https://www.threads.net/@a/post/1")
        self.assertEqual(facebook_profile_url("https://www.facebook.com/12345_67890"), "https://www.facebook.com/12345")
        self.assertEqual(profile_cache_key("ig", " @Alice "), "IG:@alice")

    def test_report_followers_are_only_filled_for_ig_and_facebook(self) -> None:
        row = {
            "post_url": "https://www.threads.com/@alice/post/ABC123",
            "username": "@alice",
            "text": "Title",
            "created_at": "2026-06-01",
            "like_count": 1,
            "reply_count": 2,
            "repost_count": 0,
            "quote_count": 0,
            "view_count": "N/A",
            "follower_count": 12345,
            "reach": "N/A",
            "status": "success",
        }

        threads_report = to_report_row(1, row, auth_mode="public", platform="THREADS")
        ig_report = to_report_row(1, row, auth_mode="public", platform="IG")
        facebook_report = to_report_row(1, row, auth_mode="public", platform="FACEBOOK")

        self.assertEqual(threads_report["粉絲團追蹤人數"], "N/A")
        self.assertEqual(ig_report["粉絲團追蹤人數"], 12345)
        self.assertEqual(facebook_report["粉絲團追蹤人數"], 12345)

    def test_report_share_is_zero_for_ig_only(self) -> None:
        row = {
            "post_url": "https://www.instagram.com/p/ABC123/",
            "username": "@alice",
            "text": "Title",
            "created_at": "2026-06-01",
            "like_count": 1,
            "reply_count": 2,
            "repost_count": 3,
            "quote_count": 4,
            "view_count": "N/A",
            "follower_count": "N/A",
            "reach": "N/A",
            "status": "success",
        }

        ig_report = to_report_row(1, row, auth_mode="public", platform="IG")
        threads_report = to_report_row(1, row, auth_mode="public", platform="THREADS")
        facebook_report = to_report_row(1, row, auth_mode="public", platform="FACEBOOK")

        self.assertEqual(ig_report["分享"], 0)
        self.assertEqual(threads_report["分享"], 7)
        self.assertEqual(facebook_report["分享"], 3)

    def test_success_report_rows_keep_unknown_metrics_as_na(self) -> None:
        threads_row = {
            "post_url": "https://www.threads.com/@alice/post/ABC123",
            "username": "@alice",
            "text": "Title",
            "created_at": "2026-06-01",
            "like_count": "N/A",
            "reply_count": "N/A",
            "repost_count": "N/A",
            "quote_count": "N/A",
            "view_count": 144,
            "follower_count": "N/A",
            "reach": "N/A",
            "status": "success",
        }
        facebook_row = dict(threads_row, post_url="https://www.facebook.com/123_456", repost_count="N/A")

        threads_report = to_report_row(1, threads_row, auth_mode="public", platform="THREADS")
        facebook_report = to_report_row(1, facebook_row, auth_mode="public", platform="FACEBOOK")

        self.assertEqual(threads_report["討論串總則數"], "N/A")
        self.assertEqual(threads_report["點閱數/按讚數"], "N/A")
        self.assertEqual(threads_report["瀏覽數"], 144)
        self.assertEqual(threads_report["分享"], "N/A")
        self.assertEqual(facebook_report["分享"], "N/A")

    def test_failed_report_rows_use_na_for_manual_backfill(self) -> None:
        row = {
            "post_url": "https://www.facebook.com/101615286547831_1339595465054527",
            "username": "麥當勞",
            "text": "Sheet title",
            "created_at": "2026-06-30",
            "like_count": 0,
            "reply_count": 0,
            "repost_count": 0,
            "quote_count": 0,
            "view_count": "N/A",
            "follower_count": "N/A",
            "reach": "N/A",
            "status": "not_found",
        }

        report = to_report_row(1, row, auth_mode="public", platform="FACEBOOK")

        self.assertEqual(report["網址"], row["post_url"])
        self.assertEqual(report["fb標題"], "Sheet title")
        self.assertEqual(report["討論串總則數"], "N/A")
        self.assertEqual(report["點閱數/按讚數"], "N/A")
        self.assertEqual(report["瀏覽數"], "N/A")
        self.assertEqual(report["分享"], "N/A")
        self.assertEqual(report["粉絲團追蹤人數"], "N/A")
        self.assertEqual(report["觸及"], "N/A")

    def test_social_follower_enrichment_retries_missing_facebook_cache(self) -> None:
        class FakeScraper:
            def scrape_many(self, urls):
                return [
                    type(
                        "Scraped",
                        (),
                        {
                            "url": urls[0],
                            "html": "",
                            "visible_text": "9.7 萬位追蹤者",
                            "network_json_blobs": [],
                        },
                    )()
                ]

        raw_rows = [{"post_url": "https://www.facebook.com/537582866282260_1443371401143730", "follower_count": "N/A"}]
        report_rows = [{"粉絲團追蹤人數": "N/A"}]
        input_items = [
            InputItem(
                url="https://www.facebook.com/537582866282260_1443371401143730",
                platform="FACEBOOK",
                author="民報",
            )
        ]
        cache = {
            profile_cache_key("FACEBOOK", "https://www.facebook.com/537582866282260"): MISSING_PROFILE_CACHE_VALUE
        }

        enrich_social_followers(raw_rows, report_rows, input_items, FakeScraper(), cache)

        self.assertEqual(raw_rows[0]["follower_count"], 97000)
        self.assertEqual(report_rows[0]["粉絲團追蹤人數"], 97000)


class ScraperHelperTests(unittest.TestCase):
    def test_scraper_url_and_response_helpers_are_stable(self) -> None:
        url = "https://www.threads.com/@alice/post/ABC123?x=1"
        self.assertEqual(_post_id_from_url(url), "ABC123")
        self.assertEqual(_profile_url_from_post_url(url), "https://www.threads.com/@alice")
        self.assertIn("www.threads.com_alice_post_ABC123", _safe_slug(url))
        self.assertTrue(_looks_like_threads_data_response("https://www.threads.com/api/graphql"))
        self.assertEqual(_parse_jsonish_response_text('for (;;);{"data":{"ok":true}}'), {"data": {"ok": True}})


class GuiPreviewTests(unittest.TestCase):
    def test_gui_url_preview_is_offline_and_platform_filtered(self) -> None:
        import gui_app  # noqa: WPS433

        payload = gui_app._preview_url_text(
            "\n".join(
                [
                    "https://www.threads.com/@alice/post/ABC123",
                    "https://www.instagram.com/p/IG123/",
                    "https://www.facebook.com/page/posts/1",
                ]
            ),
            {"THREADS", "IG"},
        )

        self.assertEqual(payload["total"], 2)
        self.assertEqual(payload["platforms"], "IG: 1, THREADS: 1")
        self.assertIn("THREADS", payload["samples"][0])

    def test_gui_xlsx_upload_preview_accepts_different_sheet_files(self) -> None:
        import gui_app  # noqa: WPS433

        with tempfile.TemporaryDirectory() as tmp:
            workbook_path = Path(tmp) / "upload.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.append(["時間", "文章標題", "作者", "來源", "網址"])
            ws.append(["2026-06-01", "Thread title", "@alice", "Threads", "https://www.threads.com/@alice/post/ABC123"])
            ws.append(["2026-06-02", "IG title", "@bob", "IG", "https://www.instagram.com/p/IG123/"])
            wb.save(workbook_path)

            client = gui_app.app.test_client()
            with workbook_path.open("rb") as uploaded:
                response = client.post(
                    "/preview",
                    data={
                        "source_mode": "xlsx",
                        "sheet_platforms": "ALL",
                        "xlsx": (uploaded, "upload.xlsx"),
                    },
                    content_type="multipart/form-data",
                )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["total"], 2)
        self.assertEqual(payload["platforms"], "IG: 1, THREADS: 1")

    def test_gui_xlsx_upload_preview_reuses_same_filename_without_stale_rows(self) -> None:
        import gui_app  # noqa: WPS433

        def make_upload(path: Path, urls: list[str]) -> None:
            wb = Workbook()
            ws = wb.active
            ws.append(["時間", "文章標題", "作者", "來源", "網址"])
            for index, url in enumerate(urls, start=1):
                ws.append([f"2026-06-{index:02d}", f"Title {index}", f"@user{index}", "Threads", url])
            wb.save(path)

        with tempfile.TemporaryDirectory() as tmp:
            first_path = Path(tmp) / "same_name.xlsx"
            second_path = Path(tmp) / "same_name_copy.xlsx"
            make_upload(first_path, [
                "https://www.threads.com/@alice/post/ABC123",
                "https://www.threads.com/@bob/post/DEF456",
            ])
            make_upload(second_path, ["https://www.threads.com/@carol/post/GHI789"])

            client = gui_app.app.test_client()
            with first_path.open("rb") as uploaded:
                first_response = client.post(
                    "/preview",
                    data={
                        "source_mode": "xlsx",
                        "sheet_platforms": "THREADS",
                        "xlsx": (uploaded, "same_name.xlsx"),
                    },
                    content_type="multipart/form-data",
                )
            with second_path.open("rb") as uploaded:
                second_response = client.post(
                    "/preview",
                    data={
                        "source_mode": "xlsx",
                        "sheet_platforms": "THREADS",
                        "xlsx": (uploaded, "same_name.xlsx"),
                    },
                    content_type="multipart/form-data",
                )

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 200)
        self.assertEqual(first_response.get_json()["total"], 2)
        self.assertEqual(second_response.get_json()["total"], 1)
        self.assertIn("carol", second_response.get_json()["samples"][0])

    def test_gui_url_preview_auto_detects_google_sheet_url(self) -> None:
        import gui_app  # noqa: WPS433

        row = type(
            "PreviewRow",
            (),
            {
                "platform": "THREADS",
                "row_number": 2,
                "url": "https://www.threads.com/@alice/post/ABC123",
            },
        )()
        client = gui_app.app.test_client()
        with patch.object(gui_app, "inspect_sheet_urls", return_value=([row], {"columns": {"title": "文章標題"}})) as mocked:
            response = client.post(
                "/preview",
                data={
                    "source_mode": "urls",
                    "sheet_platforms": "ALL",
                    "urls": "https://docs.google.com/spreadsheets/d/abc/edit?gid=123#gid=123",
                },
            )

        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["total"], 1)
        self.assertIn("Google Sheet URL detected", payload["columns"])
        mocked.assert_called_once()

    def test_gui_job_view_preserves_submitted_form_values(self) -> None:
        import gui_app  # noqa: WPS433

        class FakeThread:
            def __init__(self, *args, **kwargs):
                pass

            def start(self):
                pass

        gui_app.jobs.clear()
        try:
            client = gui_app.app.test_client()
            with patch.object(gui_app.threading, "Thread", FakeThread):
                response = client.post(
                    "/start",
                    data={
                        "source_mode": "urls",
                        "urls": "https://www.instagram.com/p/IG123/",
                        "sheet_platforms": "IG",
                        "output_name": "custom_metrics.csv",
                        "failed_output_name": "custom_failed.csv",
                        "concurrency": "2",
                        "delay": "0",
                        "retries": "0",
                        "profile_search_scrolls": "7",
                        "dry_run": "on",
                        "network_capture": "on",
                    },
                )

            self.assertEqual(response.status_code, 302)
            location = response.headers["Location"]
            job_page = client.get(location)
            html = job_page.get_data(as_text=True)

            self.assertIn('value="custom_metrics.csv"', html)
            self.assertIn('value="custom_failed.csv"', html)
            self.assertIn("https://www.instagram.com/p/IG123/", html)
            self.assertIn('<option value="IG" selected>IG</option>', html)
            self.assertIn('name="dry_run" checked', html)
            self.assertIn('name="network_capture" checked', html)
        finally:
            gui_app.jobs.clear()


if __name__ == "__main__":
    unittest.main()
